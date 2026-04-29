#!/usr/bin/env python3
"""
SIA 451 / CRBX -> Excel Parser
Erzeugt eine formatierte .xlsx-Datei aus einem SIA 451 Leistungsverzeichnis.
Benötigt: pip install openpyxl

Verwendung:
    python parse_sia451_excel.py <Eingabedatei> [Ausgabedatei.xlsx]
"""

import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from parse_sia451 import parse_sia451


# ---------------------------------------------------------------------------
# Farben & Stile
# ---------------------------------------------------------------------------
COLOR_HEADER_BG  = "1F4E79"   # dunkelblau
COLOR_HEADER_FG  = "FFFFFF"   # weiss
COLOR_SECTION_BG = "D6E4F0"   # hellblau (Abschnitts-Zeile)
COLOR_ROW_ALT    = "F2F7FB"   # sehr helles Blau (alternierende Zeilen)
COLOR_TOTAL_BG   = "1F4E79"   # wie Header (Summenzeile)
COLOR_TOTAL_FG   = "FFFFFF"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM_FMT_QTY   = '#,##0.000'
NUM_FMT_PRICE = '#,##0.00'
NUM_FMT_CHF   = '#,##0.00" CHF"'

COLUMNS = [
    ("Kapitel",        12),
    ("Abschnitt",      45),
    ("Pos-Nr",         10),
    ("Kurztext",       40),
    ("Langtext",       60),
    ("Einheit",         9),
    ("Menge",          12),
    ("Einheitspreis",  16),
    ("Gesamtpreis",    16),
    ("BKP",             8),
]
COL_NAMES  = [c[0] for c in COLUMNS]
COL_WIDTHS = [c[1] for c in COLUMNS]
COL_IDX    = {name: i + 1 for i, name in enumerate(COL_NAMES)}  # 1-basiert


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def write_header(ws) -> None:
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font      = _font(bold=True, color=COLOR_HEADER_FG)
        cell.fill      = _fill(COLOR_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22


def write_section_row(ws, row: int, section: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
    cell = ws.cell(row=row, column=1, value=section)
    cell.font      = _font(bold=True, color="1F4E79")
    cell.fill      = _fill(COLOR_SECTION_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border    = BORDER
    ws.row_dimensions[row].height = 16


def write_data_row(ws, row: int, pos: dict, alternate: bool) -> None:
    bg = COLOR_ROW_ALT if alternate else "FFFFFF"
    fill = _fill(bg)

    def cell(col_name, value, num_fmt=None, align="left"):
        c = ws.cell(row=row, column=COL_IDX[col_name], value=value)
        c.font      = _font()
        c.fill      = fill
        c.border    = BORDER
        c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
        if num_fmt:
            c.number_format = num_fmt
        return c

    cell("Kapitel",       pos["Kapitel"])
    cell("Abschnitt",     pos["Abschnitt"])
    cell("Pos-Nr",        pos["Pos-Nr"],       align="center")
    cell("Kurztext",      pos["Kurztext"])
    cell("Langtext",      pos["Langtext"])
    cell("Einheit",       pos["Einheit"],      align="center")
    cell("Menge",         pos["Menge"],        NUM_FMT_QTY,   "right")
    cell("Einheitspreis", pos["Einheitspreis"],NUM_FMT_PRICE, "right")
    cell("Gesamtpreis",   pos["Gesamtpreis"],  NUM_FMT_CHF,   "right")
    cell("BKP",           pos["BKP"],          align="center")


def write_total_row(ws, row: int, total: float) -> None:
    # Leere Zellen links
    for col in range(1, COL_IDX["Gesamtpreis"]):
        c = ws.cell(row=row, column=col, value="" if col > 1 else "TOTAL")
        c.font      = _font(bold=True, color=COLOR_TOTAL_FG)
        c.fill      = _fill(COLOR_TOTAL_BG)
        c.border    = BORDER
        c.alignment = Alignment(horizontal="right" if col == 1 else "left",
                                vertical="center")

    c = ws.cell(row=row, column=COL_IDX["Gesamtpreis"], value=total)
    c.font          = _font(bold=True, color=COLOR_TOTAL_FG)
    c.fill          = _fill(COLOR_TOTAL_BG)
    c.border        = BORDER
    c.number_format = NUM_FMT_CHF
    c.alignment     = Alignment(horizontal="right", vertical="center")

    c = ws.cell(row=row, column=COL_IDX["BKP"], value="")
    c.fill   = _fill(COLOR_TOTAL_BG)
    c.border = BORDER

    ws.row_dimensions[row].height = 18


def build_workbook(positions: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leistungsverzeichnis"

    write_header(ws)
    ws.freeze_panes = "A2"

    row = 2
    alt = False

    for pos in positions:
        write_data_row(ws, row, pos, alt)
        row += 1
        alt = not alt

    # Summen-Zeile
    total = sum(
        pos["Gesamtpreis"] for pos in positions
        if pos.get("Gesamtpreis") is not None
    )
    write_total_row(ws, row, total)

    # Auto-Filter auf Header
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    return wb


def main():
    if len(sys.argv) < 2:
        print("Verwendung: python parse_sia451_excel.py <Eingabedatei> [Ausgabedatei.xlsx]")
        sys.exit(1)

    infile  = sys.argv[1]
    outfile = sys.argv[2] if len(sys.argv) >= 3 else Path(infile).stem + ".xlsx"

    positions = parse_sia451(infile)
    if not positions:
        print("Keine Positionen gefunden.", file=sys.stderr)
        sys.exit(1)

    priced = [p for p in positions if p.get("Gesamtpreis") is not None]
    if not priced:
        print("Warnung: Keine Preise gefunden – Excel wird ohne Preisspalten erstellt.", file=sys.stderr)

    wb = build_workbook(positions)
    wb.save(outfile)

    total = sum(p["Gesamtpreis"] for p in priced)
    print(f"{len(positions)} Positionen ({len(priced)} mit Preisen), Total CHF {total:,.2f} -> {outfile}")


if __name__ == "__main__":
    main()
